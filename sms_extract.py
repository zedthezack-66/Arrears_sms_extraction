import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
from pathlib import Path
from datetime import datetime

# ── Folder structure ──────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).parent
INPUT_DIR   = SCRIPT_DIR / "input"
OUTPUT_DIR  = SCRIPT_DIR / "output"
OUTPUT_FILE = OUTPUT_DIR / "Xtenda_SMS_Extracts.xlsx"

INPUT_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# Auto-detect most recently modified .xlsx in input folder
candidates = list(INPUT_DIR.glob("*.xlsx"))
if not candidates:
    sys.exit(
        "ERROR: No .xlsx file found in the 'input' folder.\n"
        "Drop today's loanbook into the 'input' folder and re-run."
    )

INPUT_FILE = max(candidates, key=lambda f: f.stat().st_mtime)
print(f"Input file : {INPUT_FILE.name}")

# ── Sheet / style config ──────────────────────────────────────────────────────
SHEET_ORDER = [
    ("SMS_GRZ",        "GRZ"),
    ("SMS_DEFENCE",    "Defence Force"),
    ("SMS_OTHER",      "Other Employer"),
    ("SMS_OFFPAYROLL", "Off-Payroll"),
]
CAT_STYLE = {
    "GRZ":            {"hdr": "0070C0", "alt": "DDEEFF", "tab": "0070C0"},
    "Defence Force":  {"hdr": "375623", "alt": "E2EFDA", "tab": "375623"},
    "Other Employer": {"hdr": "7F5200", "alt": "FFF3CD", "tab": "FFC000"},
    "Off-Payroll":    {"hdr": "C00000", "alt": "FFE0E0", "tab": "C00000"},
}

def fill(c): return PatternFill("solid", fgColor=c)
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ── 1. Load ───────────────────────────────────────────────────────────────────
# Read without header first to detect where the real header row is (some templates
# include title rows above the column headers). If a row containing an 'Arrear'
# header is found, re-read using that row as the header.
raw = pd.read_excel(INPUT_FILE, header=None, dtype=str)
def _norm_cell(s):
    return "".join(ch.lower() for ch in str(s) if ch.isalnum())

header_row = None
for idx, row in raw.iterrows():
    for cell in row:
        if "arrear" in _norm_cell(cell):
            header_row = idx
            break
    if header_row is not None:
        break

if header_row is None:
    df = pd.read_excel(INPUT_FILE, dtype=str)
else:
    df = pd.read_excel(INPUT_FILE, header=header_row, dtype=str)

# ── 2. Locate ArrearAmount column ─────────────────────────────────────────────
# Be tolerant of minor header variations (spaces, case, punctuation).
def _norm(s):
    return "".join(ch.lower() for ch in str(s) if ch.isalnum())

_target = _norm("ArrearAmount")
arrears_col = next(
    (c for c in df.columns if _norm(c).startswith(_target) or _norm(c) == _target),
    None
)
if not arrears_col:
    available = ", ".join(map(str, df.columns))
    sys.exit(
        f"ERROR: No 'ArrearAmount' column found in '{INPUT_FILE.name}'.\n"
        f"Available columns: {available}\n"
        "Ensure the column exists and is named (or starts with) 'ArrearAmount'"
    )

print(f"Arrears col: {arrears_col}")

# ── 3. Select & clean ─────────────────────────────────────────────────────────
# Allow stable header selection across minor naming differences in the source file.
def _find_column(df, aliases, friendly_name):
    def norm(s):
        return "".join(ch.lower() for ch in str(s) if ch.isalnum())

    alias_norms = [norm(a) for a in aliases]
    for col in df.columns:
        if norm(col) in alias_norms:
            return col
    for col in df.columns:
        for alias in alias_norms:
            if alias in norm(col):
                return col

    available = ", ".join(map(str, df.columns))
    sys.exit(
        f"ERROR: No '{friendly_name}' column found in '{INPUT_FILE.name}'.\n"
        f"Available columns: {available}\n"
        f"Ensure the file contains one of: {', '.join(aliases)}"
    )

numbers_col = _find_column(df, ["NUMBERS", "NUMBER", "MOBILE", "PHONE", "MSISDN"], "NUMBERS")
customer_col = _find_column(df, ["CustomerName", "Customer Name", "ClientName", "Name"], "CustomerName")
credit_officer_col = _find_column(df, ["CREDIT OFFICER", "Credit Officer", "CreditOfficer", "Relationship Manager"], "CREDIT OFFICER")
employer_cat_col = _find_column(df, ["EmployerCat", "Employer Category", "EmployerCat", "Employer"], "EmployerCat")

df_out = df[[numbers_col, customer_col, arrears_col, credit_officer_col, employer_cat_col]].copy()
df_out.columns = ["NUMBER", "CUSTOMER NAME", "ArrearAmount", "CREDIT OFFICER", "EmployerCat"]

df_out["NUMBER"] = df_out["NUMBER"].astype(str)
df_out["CUSTOMER NAME"] = df_out["CUSTOMER NAME"].astype(str)
df_out["ArrearAmount"] = pd.to_numeric(df_out["ArrearAmount"], errors="coerce").round(0)

df_out = df_out[df_out["NUMBER"].notna() & (df_out["NUMBER"].str.strip() != "")]
df_out = df_out[df_out["CUSTOMER NAME"].notna() & (df_out["CUSTOMER NAME"].str.strip() != "")]
df_out = df_out[df_out["ArrearAmount"].notna() & (df_out["ArrearAmount"] != 0)]
print(f"Clean rows : {len(df_out):,}")

# ── 4. Build output workbook ──────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)
summary = {}

for sheet_name, cat in SHEET_ORDER:
    subset = df_out[df_out["EmployerCat"] == cat][
        ["NUMBER", "CUSTOMER NAME", "ArrearAmount", "CREDIT OFFICER"]
    ].reset_index(drop=True)

    summary[sheet_name] = len(subset)
    style = CAT_STYLE[cat]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = style["tab"]

    # Title banner
    ws.merge_cells("A1:D1")
    ws["A1"] = f"XTENDA FINANCIAL SERVICES  —  {sheet_name}  |  {cat}  |  {INPUT_FILE.stem}"
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws["A1"].fill = fill(style["hdr"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws["F1"] = "Records:"
    ws["F1"].font = Font(name="Arial", bold=True, size=10, color=style["hdr"])
    ws["G1"] = len(subset)
    ws["G1"].font = Font(name="Arial", bold=True, size=10, color=style["hdr"])
    ws["G1"].number_format = "#,##0"

    # Column headers
    for ci, h in enumerate(["NUMBER", "CUSTOMER NAME", "ArrearAmount", "CREDIT OFFICER"], start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = fill(style["hdr"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    ws.row_dimensions[2].height = 20

    # Data rows
    for ri, row in subset.iterrows():
        er = ri + 3
        for ci, val in enumerate(
            [row["NUMBER"], row["CUSTOMER NAME"], row["ArrearAmount"], row["CREDIT OFFICER"]],
            start=1
        ):
            cell = ws.cell(row=er, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center")
            if ci == 3:
                cell.number_format = "#,##0"
            if ri % 2 == 0:
                cell.fill = fill(style["alt"])
        ws.row_dimensions[er].height = 18

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 26
    ws.auto_filter.ref = "A2:D2"
    ws.freeze_panes = "A3"

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet(title="SUMMARY", index=0)
ws_sum.sheet_properties.tabColor = "1F4E79"

ws_sum.merge_cells("A1:D1")
ws_sum["A1"] = "XTENDA FINANCIAL SERVICES — SMS EXTRACT SUMMARY"
ws_sum["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
ws_sum["A1"].fill = fill("1F4E79")
ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 32

for r, (lbl, val) in enumerate([
    ("Generated:",      datetime.now().strftime("%d/%m/%Y %H:%M")),
    ("Source File:",    INPUT_FILE.name),
    ("Arrears Column:", arrears_col),
], start=2):
    ws_sum.cell(row=r, column=1, value=lbl).font = Font(name="Arial", bold=True, size=10)
    ws_sum.cell(row=r, column=2, value=val).font  = Font(name="Arial", size=10)

ws_sum.row_dimensions[5].height = 10
for ci, h in enumerate(["Sheet", "Category", "Records"], start=1):
    cell = ws_sum.cell(row=6, column=ci, value=h)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = fill("1F4E79")
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border()
ws_sum.row_dimensions[6].height = 20

total = 0
for ri, (sheet_name, cat) in enumerate(SHEET_ORDER, start=7):
    count = summary[sheet_name]
    total += count
    style = CAT_STYLE[cat]
    for ci, val in enumerate([sheet_name, cat, count], start=1):
        cell = ws_sum.cell(row=ri, column=ci, value=val)
        cell.font = Font(name="Arial", size=10, bold=(ci == 1))
        cell.fill = fill(style["alt"])
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center" if ci != 2 else "left", vertical="center")
        if ci == 3:
            cell.number_format = "#,##0"
    ws_sum.row_dimensions[ri].height = 20

total_row = 7 + len(SHEET_ORDER)
for ci, val in enumerate(["TOTAL", "", total], start=1):
    cell = ws_sum.cell(row=total_row, column=ci, value=val)
    cell.font = Font(name="Arial", bold=True, size=10)
    cell.fill = fill("D9D9D9")
    cell.border = thin_border()
    if ci == 3:
        cell.number_format = "#,##0"
ws_sum.row_dimensions[total_row].height = 20

ws_sum.column_dimensions["A"].width = 18
ws_sum.column_dimensions["B"].width = 28
ws_sum.column_dimensions["C"].width = 14

wb.save(OUTPUT_FILE)

print(f"\n{'─'*40}")
for sname, _ in SHEET_ORDER:
    print(f"  {sname:20s}  {summary[sname]:>5,} records")
print(f"  {'TOTAL':20s}  {total:>5,} records")
print(f"{'─'*40}")
print(f"\nSaved → output/{OUTPUT_FILE.name}")
