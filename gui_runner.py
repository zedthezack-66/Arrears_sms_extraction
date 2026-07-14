import threading
import shutil
import sys
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Processing function adapted from sms_extract.py
CAT_STYLE = {
    "GRZ":            {"hdr": "0070C0", "alt": "DDEEFF", "tab": "0070C0"},
    "Defence Force":  {"hdr": "375623", "alt": "E2EFDA", "tab": "375623"},
    "Other Employer": {"hdr": "7F5200", "alt": "FFF3CD", "tab": "FFC000"},
    "Off-Payroll":    {"hdr": "C00000", "alt": "FFE0E0", "tab": "C00000"},
}
SHEET_ORDER = [
    ("SMS_GRZ",        "GRZ"),
    ("SMS_DEFENCE",    "Defence Force"),
    ("SMS_OTHER",      "Other Employer"),
    ("SMS_OFFPAYROLL", "Off-Payroll"),
]

def fill(c): return PatternFill("solid", fgColor=c)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _norm_cell(s):
    return "".join(ch.lower() for ch in str(s) if ch.isalnum())


def _norm(s):
    return "".join(ch.lower() for ch in str(s) if ch.isalnum())


def _find_column(df, aliases, friendly_name):
    def norm(s):
        return "".join(ch.lower() for ch in str(s) if ch.isalnum())

    alias_norms = [norm(a) for a in aliases]
    for col in df.columns:
        if norm(col) in alias_norms:
            return col
    for col in df.columns:
        for alias in alias_norms:
            if alias in norm(col):
                return col

    available = ", ".join(map(str, df.columns))
    raise ValueError(
        f"ERROR: No '{friendly_name}' column found in source. Available columns: {available}."
    )


def process_file(input_path, base_dir=None, log_fn=print):
    input_path = Path(input_path)
    if base_dir is None:
        base_dir = Path.cwd()
    base_dir = Path(base_dir)
    output_dir = base_dir / "output"
    output_dir.mkdir(exist_ok=True)
    output_file = output_dir / "Xtenda_SMS_Extracts.xlsx"

    log_fn(f"Input file : {input_path.name}")

    # Read
    raw = pd.read_excel(input_path, header=None, dtype=str)
    header_row = None
    for idx, row in raw.iterrows():
        for cell in row:
            if "arrear" in _norm_cell(cell):
                header_row = idx
                break
        if header_row is not None:
            break

    if header_row is None:
        df = pd.read_excel(input_path, dtype=str)
    else:
        df = pd.read_excel(input_path, header=header_row, dtype=str)

    _target = _norm("ArrearAmount")
    arrears_col = next(
        (c for c in df.columns if _norm(c).startswith(_target) or _norm(c) == _target),
        None
    )
    if not arrears_col:
        available = ", ".join(map(str, df.columns))
        raise ValueError(
            f"ERROR: No 'ArrearAmount' column found in '{input_path.name}'. Available columns: {available}"
        )
    log_fn(f"Arrears col: {arrears_col}")

    numbers_col = _find_column(df, ["NUMBERS", "NUMBER", "MOBILE", "PHONE", "MSISDN"], "NUMBERS")
    customer_col = _find_column(df, ["CustomerName", "Customer Name", "ClientName", "Name"], "CustomerName")
    credit_officer_col = _find_column(df, ["CREDIT OFFICER", "Credit Officer", "CreditOfficer", "Relationship Manager"], "CREDIT OFFICER")
    employer_cat_col = _find_column(df, ["EmployerCat", "Employer Category", "EmployerCat", "Employer"], "EmployerCat")

    df_out = df[[numbers_col, customer_col, arrears_col, credit_officer_col, employer_cat_col]].copy()
    df_out.columns = ["NUMBER", "CUSTOMER NAME", "ArrearAmount", "CREDIT OFFICER", "EmployerCat"]

    df_out["NUMBER"] = df_out["NUMBER"].astype(str)
    df_out["CUSTOMER NAME"] = df_out["CUSTOMER NAME"].astype(str)
    df_out["ArrearAmount"] = pd.to_numeric(df_out["ArrearAmount"], errors="coerce").round(0)

    df_out = df_out[df_out["NUMBER"].notna() & (df_out["NUMBER"].str.strip() != "")]
    df_out = df_out[df_out["CUSTOMER NAME"].notna() & (df_out["CUSTOMER NAME"].str.strip() != "")]
    df_out = df_out[df_out["ArrearAmount"].notna() & (df_out["ArrearAmount"] != 0)]
    log_fn(f"Clean rows : {len(df_out):,}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    summary = {}

    for sheet_name, cat in SHEET_ORDER:
        subset = df_out[df_out["EmployerCat"] == cat][["NUMBER", "CUSTOMER NAME", "ArrearAmount", "CREDIT OFFICER"]].reset_index(drop=True)
        summary[sheet_name] = len(subset)
        style = CAT_STYLE[cat]
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = style["tab"]

        ws.merge_cells("A1:D1")
        ws["A1"] = f"XTENDA FINANCIAL SERVICES  —  {sheet_name}  |  {cat}  |  {input_path.stem}"
        ws["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ws["A1"].fill = fill(style["hdr"])
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        ws["F1"] = "Records:"
        ws["F1"].font = Font(name="Arial", bold=True, size=10, color=style["hdr"])
        ws["G1"] = len(subset)
        ws["G1"].font = Font(name="Arial", bold=True, size=10, color=style["hdr"])
        ws["G1"].number_format = "#,##0"

        for ci, h in enumerate(["NUMBER", "CUSTOMER NAME", "ArrearAmount", "CREDIT OFFICER"], start=1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill = fill(style["hdr"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()
        ws.row_dimensions[2].height = 20

        for ri, row in subset.iterrows():
            er = ri + 3
            for ci, val in enumerate([row["NUMBER"], row["CUSTOMER NAME"], row["ArrearAmount"], row["CREDIT OFFICER"]], start=1):
                cell = ws.cell(row=er, column=ci, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border()
                cell.alignment = Alignment(vertical="center")
                if ci == 3:
                    cell.number_format = "#,##0"
                if ri % 2 == 0:
                    cell.fill = fill(style["alt"])
            ws.row_dimensions[er].height = 18

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 26
        ws.auto_filter.ref = "A2:D2"
        ws.freeze_panes = "A3"

    ws_sum = wb.create_sheet(title="SUMMARY", index=0)
    ws_sum.sheet_properties.tabColor = "1F4E79"

    ws_sum.merge_cells("A1:D1")
    ws_sum["A1"] = "XTENDA FINANCIAL SERVICES — SMS EXTRACT SUMMARY"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws_sum["A1"].fill = fill("1F4E79")
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 32

    for r, (lbl, val) in enumerate([
        ("Generated:",      datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Source File:",    input_path.name),
        ("Arrears Column:", arrears_col),
    ], start=2):
        ws_sum.cell(row=r, column=1, value=lbl).font = Font(name="Arial", bold=True, size=10)
        ws_sum.cell(row=r, column=2, value=val).font  = Font(name="Arial", size=10)

    ws_sum.row_dimensions[5].height = 10
    for ci, h in enumerate(["Sheet", "Category", "Records"], start=1):
        cell = ws_sum.cell(row=6, column=ci, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = fill("1F4E79")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()
    ws_sum.row_dimensions[6].height = 20

    total = 0
    for ri, (sheet_name, cat) in enumerate(SHEET_ORDER, start=7):
        count = summary[sheet_name]
        total += count
        style = CAT_STYLE[cat]
        for ci, val in enumerate([sheet_name, cat, count], start=1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10, bold=(ci == 1))
            cell.fill = fill(style["alt"])
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if ci != 2 else "left", vertical="center")
            if ci == 3:
                cell.number_format = "#,##0"
        ws_sum.row_dimensions[ri].height = 20

    total_row = 7 + len(SHEET_ORDER)
    for ci, val in enumerate(["TOTAL", "", total], start=1):
        cell = ws_sum.cell(row=total_row, column=ci, value=val)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.fill = fill("D9D9D9")
        cell.border = thin_border()
        if ci == 3:
            cell.number_format = "#,##0"
    ws_sum.row_dimensions[total_row].height = 20

    ws_sum.column_dimensions["A"].width = 18
    ws_sum.column_dimensions["B"].width = 28
    ws_sum.column_dimensions["C"].width = 14

    wb.save(output_file)

    log_fn(f"Saved → {output_dir.name}/{output_file.name}")
    return str(output_file)


# Simple Tkinter GUI
class App:
    def __init__(self, root):
        self.root = root
        root.title("XTENDA SMS Extractor")
        root.geometry("640x360")

        frm = tk.Frame(root)
        frm.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.path_var = tk.StringVar()
        tk.Button(frm, text="Choose Excel (.xlsx)", command=self.choose).pack(anchor=tk.W)
        tk.Label(frm, textvariable=self.path_var, wraplength=600).pack(anchor=tk.W, pady=(4,8))

        self.run_btn = tk.Button(frm, text="Run", state=tk.DISABLED, command=self.start)
        self.run_btn.pack(anchor=tk.W)

        tk.Label(frm, text="Log:").pack(anchor=tk.W, pady=(10,0))
        self.log = scrolledtext.ScrolledText(frm, height=12)
        self.log.pack(fill=tk.BOTH, expand=True)

    def choose(self):
        p = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if p:
            self.path_var.set(p)
            self.run_btn.config(state=tk.NORMAL)

    def append(self, s):
        self.log.insert(tk.END, s + "\n")
        self.log.see(tk.END)

    def start(self):
        inp = self.path_var.get()
        if not inp:
            messagebox.showerror("No file", "Please select an Excel file first.")
            return
        self.run_btn.config(state=tk.DISABLED)
        threading.Thread(target=self._run, args=(inp,), daemon=True).start()

    def _run(self, inp):
        try:
            self.append(f"Starting at {datetime.now().strftime('%H:%M:%S')}")
            out = process_file(inp, base_dir=Path.cwd(), log_fn=self.append)
            self.append(f"Completed: {out}")
            messagebox.showinfo("Done", f"Output saved to:\n{out}")
        except Exception as e:
            self.append(f"ERROR: {e}")
            messagebox.showerror("Error", str(e))
        finally:
            self.run_btn.config(state=tk.NORMAL)


def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()


if __name__ == '__main__':
    main()
